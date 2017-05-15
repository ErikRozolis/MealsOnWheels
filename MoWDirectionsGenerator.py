import googlemaps, json, io, jdcal
import openpyxl, os, sys
from openpyxl.styles import Alignment
from openpyxl import Workbook
from HTMLParser import HTMLParser
from datetime import datetime
import Tkinter, Tkconstants, tkFileDialog, os
from Tkinter import *

class Route(object):
    routeName = ""
    routeStops = []

    def __init__(self, routeName):
        self.routeName = routeName
        self.routeStops = []

class RouteStop(object):
    number = ""
    clientName = ""
    address = ""
    phone = ""
    hotMeals = 0
    coldMeals = 0
    deliveryInstructions = ""
    specialInstructions = ""
    sunday = False
    monday = False
    tuesday = False
    wednesday = False
    thursday = False
    friday = False
    saturday = False

    # The class "constructor" - It's actually an initializer
    def __init__(self, number, clientName, address, phone):
        self.number = number
        self.clientName = clientName
        self.phone = phone
        self.address = address

class MLStripper(HTMLParser):
    def __init__(self):
        self.reset()
        self.fed = []
    def handle_starttag(self, tag, attrs):
        if('div' in tag):
            self.fed.append("\n")
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ''.join(self.fed)

def strip_tags(html):
    s = MLStripper()
    s.feed(html)
    return s.get_data()

def compileWorkBooks(frame):
    wbRead = openpyxl.load_workbook(frame.selectedDir)
    wbWrite = Workbook()
    wbWrite.active.cell(column=1,row=1,value="For each route, please see the individual sheets below")

    allStopsList = []
    routeList = {}

    iterSheet = iter(wbRead.active)
    next(iterSheet)
    for row in iterSheet:
        allStopsList.append(RouteStop(row[0].value, row[1].value, row[2].value, row[3].value))
    wbRead.close()

    for routestop in allStopsList:
        if routestop.number[0] not in routeList:
            routeList[str(routestop.number[0])]=Route(routestop.number[0])
        routeList[routestop.number[0]].routeStops.append(routestop)
    print routeList
    for route in routeList:
        #gmaps = googlemaps.Client(key="AIzaSyD7KkieqvubRP-vng1PcYToF8fzxBDS5cU")
        gmaps = googlemaps.Client(key=frame.key.get())
        routeAddresses = ""
        for stop in routeList[route].routeStops:
            routeAddresses = "|".join([routeAddresses, stop.address])
        directionsList = gmaps.directions(frame.headquarters.get(), frame.headquarters.get(),
                                          mode="driving", waypoints=routeAddresses, optimize_waypoints=True,
                                          departure_time=datetime.now())
        directions = directionsList[0]
        print directions['waypoint_order']
        row = 1
        counter = 1
        max_length = 0
        wbWrite.create_sheet(route)
        for orderedNum in directions['waypoint_order']:
            print routeList[route].routeStops[orderedNum].number
            wbWrite[route].cell(column=1, row=row, value=routeList[route].routeStops[orderedNum].number)
            wbWrite[route].cell(column=2, row=row, value=routeList[route].routeStops[orderedNum].clientName)
            wbWrite[route].cell(column=3, row=row, value=routeList[route].routeStops[orderedNum].phone)
            wbWrite[route].cell(column=4, row=row, value=routeList[route].routeStops[orderedNum].address)
            row+=1
        for leg in directions['legs']:
            instructionString = ""
            for steps in leg['steps']:
                instructionChunk = strip_tags(json.dumps(steps['html_instructions'], ensure_ascii=False))[1:-1] + "\n"
                instructionString += instructionChunk
                try:
                    if (len(instructionChunk) > max_length):
                        max_length = len(instructionChunk)
                except:
                    pass
                wbWrite[route].cell(column=5, row=counter, value=instructionString[:-1])
            wbWrite[route].cell(column=5, row=counter).alignment = Alignment(wrapText=True)
            counter += 1
        adjusted_directions_width = max_length

        wbWrite[route].column_dimensions['E'].width = adjusted_directions_width
    wbWrite.save("OutputWorkbook.xlsx")

class App(Tkinter.Frame):
    def __init__(self, master):
        self.fileCount=0
        Tkinter.Frame.__init__(self, root)

        button_opt = {'fill':Tkconstants.BOTH, 'padx':5, 'pady':5}

        self.displayBox = Tkinter.Text()
        self.displayBox.pack()
        self.key = Entry(root)
        self.key.insert(END, "AIzaSyD7KkieqvubRP-vng1PcYToF8fzxBDS5cU")
        self.key.pack()
        self.headquarters = Entry(root)
        self.headquarters.insert(END, "424 College St, Beloit, WI 53511")
        self.headquarters.pack()

        self.dir_opt = options = {}
        options['initialdir'] = os.path.dirname(os.path.realpath(sys.argv[0]))
        options['parent'] = root
        options['title'] = 'Select a file to parse'

        Tkinter.Button(self, text='Select File to Parse', command=self.askdirectory).pack(**button_opt)
    def askdirectory(self):
        self.fileCount=0
        self.selectedDir = tkFileDialog.askopenfilename(**self.dir_opt)
        self.displayBox.insert(Tkinter.END, self.selectedDir + " selected")
        compileWorkBooks(self)

if __name__=='__main__':
    root = Tkinter.Tk()
    App(root).pack()
    root.mainloop()