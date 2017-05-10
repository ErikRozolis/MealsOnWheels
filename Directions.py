import googlemaps, json, io, jdcal, openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
from HTMLParser import HTMLParser
from datetime import datetime

htmlParser = HTMLParser()

addresses = ""

class MLStripper(HTMLParser):
    def __init__(self):
        self.reset()
        self.fed = []
    def handle_starttag(self, tag, attrs):
        if('div' in tag):
            self.fed.append("\n")
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ''.join(self.fed)

def strip_tags(html):
    s = MLStripper()
    s.feed(html)
    return s.get_data()

wb = openpyxl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']

wbwrite = Workbook()
wswrite = wbwrite.active
wswrite.title = "Output"
# wswrite.row_dimensions[1].height = '30'
wswrite.column_dimensions['E'].width =50
wswrite.column_dimensions['B'].width = 50

header = [cell.value for cell in sheet[1]]

itersheet = iter(sheet)

next(itersheet)

for row in itersheet:
	number = row[0].value
	address = row[1].value
	hotmeals = row[2].value
	coldmeals = row[3].value
	addresses += str(row[1].value + " Beloit WI " + "|")

gmaps = googlemaps.Client(key='AIzaSyD7KkieqvubRP-vng1PcYToF8fzxBDS5cU')



directionsList = gmaps.directions("424 College St, Beloit, WI 53511","424 College St, Beloit, WI 53511",
                                     mode="driving", waypoints=addresses, optimize_waypoints=True, departure_time=datetime.now())
# with io.open("Fulldirections.json", "w", encoding='utf-8') as file:
	# file.write(unicode(json.dumps(directionsList, ensure_ascii=False)))
directions = directionsList[0]
print directions['waypoint_order']
for row in directions['waypoint_order']:
	number = sheet[row+2][0].value
	address = sheet[row+2][1].value
	hotmeals = sheet[row+2][2].value
	coldmeals = sheet[row+2][3].value
	wswrite.cell(column = 1, row = row+1, value = number)
	wswrite.cell(column = 2, row = row+1, value = address)
	wswrite.cell(column = 3, row = row+1, value = hotmeals)
	wswrite.cell(column = 4, row = row+1, value = coldmeals)
	
counter = 0
with io.open("outputfile.json", "w", encoding='utf-8') as outputfile:
	for leg in directions['legs']:
		instructionString = ""
		for steps in leg['steps']:
			instructionString += strip_tags(json.dumps(steps['html_instructions'], ensure_ascii=False))[1:-1] + "\n"
		# finalIndex = instructionString.find("Destination")
		# additionalInstructionString = "\n" + instructionString[finalIndex:-2]
		wswrite.cell(column = 5, row = counter + 1, value = instructionString[:-1])
		wswrite.cell(column=5, row = counter+1).alignment = Alignment(wrapText=True)
		counter+=1
		
wbwrite.save('outputWorkbook.xlsx')